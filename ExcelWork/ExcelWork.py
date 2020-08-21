from openpyxl import load_workbook 
workbook = load_workbook(filename = "RJ目录.xlsx")
print(workbook.sheetnames) 
workbook.sheetnames 
sheet = workbook.active
print(sheet)


cell1 = sheet["A1"] 
cell2 = sheet["N11"] 
print(cell1.value, cell2.value)
cell3 = sheet.cell(row = 1,column = 1) 
cell4 = sheet.cell(row = 11,column = 3)
print(cell3.value, cell4.value)
print(cell1.value, cell1.row, cell1.column, cell1.coordinate) 
print(cell2.value, cell2.row, cell2.column, cell2.coordinate) 
""" .row 获取某个格子的行数； 
.columns 获取某个格子的列数； 
.corordinate 获取某个格子的坐标； """
""" workbook.active 打开激活的表格； 
sheet["A1"] 获取A1格子的数据； 
cell.value 获取格子中的值； """
